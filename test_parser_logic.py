import openpyxl
import os

# Путь к файлу (который мы проверяли рентгеном)
FILE_PATH = "downloads/ФВМ/4 курс/КММБ ММВЯ УМПД (22.12.2025-27.12.2025).xlsx"

# Та самая функция из Строгого парсера
def get_value_from_merged(sheet, row, col):
    cell = sheet.cell(row, col)
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            # Если ячейка объединена, возвращаем значение ГЛАВНОЙ ячейки
            main_val = sheet.cell(merged.min_row, merged.min_col).value
            return f"✅ MERGED -> {main_val}"
    # Если не объединена
    return f"❌ NOT MERGED (Value: {cell.value})"

def test():
    if not os.path.exists(FILE_PATH):
        print(f"Файл не найден: {FILE_PATH}")
        return

    print(f"📂 Открываю: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    sheet = wb.active

    # Мы знаем из рентгена, что:
    # Строка = 29
    # Группа КММБ (Колонка 3) - Главная
    # Группа ММВЯ (Колонка 4) - Пустая, но должна быть объединена
    
    row = 29
    
    print(f"\n--- ТЕСТ СТРОКИ {row} ---")
    
    # Проверяем колонку 3 (КММБ)
    val3 = get_value_from_merged(sheet, row, 3)
    print(f"Col 3 (КММБ): {val3}")
    
    # Проверяем колонку 4 (ММВЯ) - САМОЕ ВАЖНОЕ
    val4 = get_value_from_merged(sheet, row, 4)
    print(f"Col 4 (ММВЯ): {val4}")

if __name__ == "__main__":
    test()