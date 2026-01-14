import openpyxl
import os
import sys

# Путь к файлу 4 курса ФВМ
SEARCH_DIR = "downloads/ФВМ/4 курс"
DATE_MARKER = "22.12" # Чтобы найти нужный файл

def xray_check():
    # 1. Ищем файл
    target_file = None
    if os.path.exists(SEARCH_DIR):
        for f in os.listdir(SEARCH_DIR):
            if f.endswith(".xlsx") and DATE_MARKER in f:
                target_file = os.path.join(SEARCH_DIR, f)
                break
    
    if not target_file:
        print("❌ Файл Excel не найден!")
        return

    print(f"🩻 Рентген файла: {target_file}")
    wb = openpyxl.load_workbook(target_file)
    sheet = wb.active

    # 2. Ищем строку с 26.12 (Пятница)
    target_row = None
    for row in range(1, 40):
        val = str(sheet.cell(row, 1).value)
        if "26.12" in val:
            target_row = row
            print(f"📍 Нашел дату '26.12' на строке {target_row}")
            break
            
    if not target_row:
        print("❌ Не нашел дату 26.12 в файле!")
        return

    # 3. Проверяем колонки групп (обычно это 3, 4, 5, 6, 7)
    # 1=День, 2=Время, 3=Группа1, 4=Группа2...
    print("\n--- ПРОВЕРКА ЯЧЕЕК ---")
    for col in range(3, 8):
        cell = sheet.cell(target_row, col)
        
        # Проверка на объединение
        is_merged = False
        merged_range_str = "Нет"
        
        for rng in sheet.merged_cells.ranges:
            if cell.coordinate in rng:
                is_merged = True
                merged_range_str = str(rng)
                break
        
        value = str(cell.value).strip() if cell.value else "[ПУСТО]"
        
        print(f"Column {col} ({cell.coordinate}):")
        print(f"   📝 Значение: {value[:20]}...")
        print(f"   🔗 Объединена? {'✅ ДА' if is_merged else '❌ НЕТ'}")
        if is_merged:
            print(f"      Диапазон: {merged_range_str}")
        print("-" * 20)

if __name__ == "__main__":
    xray_check()