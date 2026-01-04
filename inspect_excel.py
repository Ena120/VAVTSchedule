import openpyxl
import os

def inspect():
    # 1. Ищем файл Excel для ФВМ (где была проблема)
    target_file = None
    for root, dirs, files in os.walk("downloads"):
        for file in files:
            # Ищем файл этой недели
            if file.endswith(".xlsx") and "22.12" in file:
                target_file = os.path.join(root, file)
                break
        if target_file: break
    
    if not target_file:
        print("❌ Файл не найден! Убедись, что run_update.py отработал.")
        return

    print(f"🔎 Анализирую: {target_file}")
    wb = openpyxl.load_workbook(target_file, data_only=True)
    sheet = wb.active

    # 2. Выводим строки (примерно с 20 по 35, где пятница)
    print("\n--- СРЕЗ ТАБЛИЦЫ (Строки 20-35) ---")
    for i in range(1, sheet.max_column + 1):
        print(f"Col {i}", end="\t")
    print("\n")

    # Ищем строку с "26.12"
    start_row = 1
    for row in range(1, sheet.max_row + 1):
        val = str(sheet.cell(row, 1).value)
        if "26.12" in val:
            start_row = row
            print(f"📍 Нашел Пятницу 26.12 на строке {row}")
            break
            
    # Печатаем 15 строк начиная с пятницы
    for row in range(start_row, start_row + 15):
        row_data = []
        for col in range(1, 6): # Первые 5 колонок
            cell = sheet.cell(row, col)
            val = str(cell.value).strip() if cell.value else "."
            # Если ячейка объединена, пометим это
            is_merged = False
            for rng in sheet.merged_cells.ranges:
                if cell.coordinate in rng:
                    is_merged = True
                    break
            
            marker = "[M]" if is_merged else ""
            row_data.append(f"{val[:15]}{marker}") # Обрезаем длинный текст
            
        print(f"Row {row}: {row_data}")

if __name__ == "__main__":
    inspect()