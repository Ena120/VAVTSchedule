import openpyxl
import re

def get_value_from_merged(sheet, row, col):
    """
    Берет значение с учетом объединения ячеек.
    (Проверено тестом: работает с файлами Adobe)
    """
    cell = sheet.cell(row, col)
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value

def clean_text(text):
    if not text or str(text) == "None": 
        return ""
    return str(text).strip().replace('\n', ' ')

def parse_schedule(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    schedule_data = []
    
    # --- 1. Поиск шапки ---
    header_row_index = -1
    group_columns = {}
    group_pattern = re.compile(r'^[А-ЯЁ]\d{2}.*') 

    for row_num in range(1, 30):
        for col_num in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row_num, column=col_num).value
            if val and isinstance(val, str):
                val = val.replace('\n', '')
                if group_pattern.match(val.strip()):
                    header_row_index = row_num
                    for c in range(1, sheet.max_column + 1):
                        g_val = sheet.cell(row=row_num, column=c).value
                        if g_val and isinstance(g_val, str) and len(g_val) > 2:
                             group_columns[c] = g_val.strip().replace('\n', '')
                    break
        if header_row_index != -1: break
            
    if header_row_index == -1:
        print(f"⚠️ Шапка не найдена: {file_path}")
        return []

    # --- 2. Чтение данных ---
    current_day = None
    
    for row_num in range(header_row_index + 1, sheet.max_row + 1):
        # А. День недели
        day_val = get_value_from_merged(sheet, row_num, 1)
        day_str = clean_text(day_val)
        
        if day_str:
            # Фильтр мусора в подвале
            if len(day_str) < 15 and "корпус" not in day_str.lower():
                current_day = day_str
        
        if not current_day: continue

        # Б. Время
        col2_val = get_value_from_merged(sheet, row_num, 2)
        col2_str = clean_text(col2_val)
        
        is_time = False
        if len(col2_str) < 15 and any(c.isdigit() for c in col2_str):
            is_time = True
        
        if is_time:
            base_time = col2_str
            subject_prefix = ""
        else:
            base_time = "🕒 См. описание"
            subject_prefix = f"[{col2_str}] " if col2_str else ""

        # Г. Проход по группам
        for col_idx, group_name in group_columns.items():
            # Используем ту же функцию, что и в тесте!
            raw_val = get_value_from_merged(sheet, row_num, col_idx)
            subject_text = clean_text(raw_val)
            
            # Если после проверки Merge там все равно пусто - значит пары нет.
            if not subject_text:
                continue

            final_time = base_time
            if not is_time:
                time_match = re.search(r'(\d{1,2}[:.]\d{2})', subject_text)
                if time_match:
                    final_time = time_match.group(1)

            full_subject = f"{subject_prefix}{subject_text}"

            schedule_data.append({
                "day": current_day,
                "time": final_time,
                "group": group_name,
                "subject_raw": full_subject
            })

    return schedule_data